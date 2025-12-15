"""Script to replay scenarios via HTTP APIs per custom requirements.

Usage examples (PowerShell):

- Default (uses built-in presets):
    `python usecase_hot.py`

- Provide a JSON/YAML config with your requirements:
    `python usecase_hot.py --config .\\my_scenario.json`

- Preview without sending requests:
    `python usecase_hot.py --dry-run`

The config file may define `baseUrl`, `hyperparams`, `rooms`, and `timeline`.
See the bottom of this file or README for a sample.
"""
from __future__ import annotations

import time
from typing import Any, Dict, Iterable, List, Optional

import argparse
import json
from pathlib import Path
try:
    import yaml  # optional, for YAML config support
except Exception:
    yaml = None

import requests
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import box
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

BASE_URL = "http://localhost:8000"
SESSION = requests.Session()
DRY_RUN = False
CONSOLE = Console()
SNAPSHOT_ROWS: List[Dict[str, Any]] = []

# ---------------------------------------------------------------------------
# 1) 管理员全局参数：如果某个键设为 None，会落回当前配置文件的默认值。
HYPERPARAM_OVERRIDES: Dict[str, float] = {
    # 调度 + 节流
    "maxConcurrent": 3,
    "timeSliceSeconds": 120,
    "changeTempMs": 1000,
    "autoRestartThreshold": 0.5,
    # 温控：来自需求表格（制热 18-25℃、缺省 23℃、不同风速的升温速率）
    "coolRangeMin": 18.0,
    "coolRangeMax": 28.0,
    "heatRangeMin": 18.0,
    "heatRangeMax": 25.0,
    "idleDriftPerMin": 0.5,
    "midDeltaPerMin": 0.5,  # 1℃/2min
    "highMultiplier": 2,  # -> 1℃/1min
    "lowMultiplier": 2/3,  # -> 1℃/3min
    "defaultTarget": 23.0,
    # 计费：1 元/1℃，不同风速对应每分钟单价
    "pricePerUnit": 1.0,
    "rateHighUnitPerMin": 1.0,
    "rateMidUnitPerMin": 0.5,
    "rateLowUnitPerMin": 1.0 / 3.0,
    # 住宿默认单价（单房自定义仍通过 open_room 设置）
    "ratePerNight": 150.0,
    # 时钟倍率：ratio=60 代表 1 分钟的业务时间约等于 1 秒真实时间
    "clockRatio": 60.0,
}

# ---------------------------------------------------------------------------
# 2) 房间初始化：来自「房间初始温度」表。可增删房间、修改初温和房价。
ROOM_PRESETS: List[Dict[str, Any]] = [
    {"roomId": "1", "initialTemp": 10.0, "ratePerNight": 100.0},
    {"roomId": "2", "initialTemp": 15.0, "ratePerNight": 125.0},
    {"roomId": "3", "initialTemp": 18.0, "ratePerNight": 150.0},
    {"roomId": "4", "initialTemp": 12.0, "ratePerNight": 200.0},
    {"roomId": "5", "initialTemp": 14.0, "ratePerNight": 100.0},
]

# ---------------------------------------------------------------------------
# 3) 时间轴：根据 Excel 中每分钟的操作填写。以下内容由截图推断，可根据
#    实际需要自由修改 / 扩展。
#    每项 action = {"roomId", "type", "payload"}，type 取值：
#      - power_on  -> POST /rooms/{id}/ac/power-on
#      - power_off -> POST /rooms/{id}/ac/power-off
#      - change_temp -> POST /rooms/{id}/ac/change-temp
#      - change_speed -> POST /rooms/{id}/ac/change-speed
TIMELINE: Dict[int, List[Dict[str, Any]]] = {
    1: [
        {"roomId": "1", "type": "power_on"},
    ],
    2: [
        {"roomId": "1", "type": "change_temp", "payload": {"targetTemp": 24.0}},
        {"roomId": "2", "type": "power_on"}
    ],
    3: [
        {"roomId": "3", "type": "power_on"},
    ],
    4: [
        {"roomId": "2", "type": "change_temp", "payload": {"targetTemp": 25.0}},
        {"roomId": "4", "type": "power_on"},
        {"roomId": "5", "type": "power_on"},
    ],
    5: [
        {"roomId":"3", "type": "change_temp", "payload": {"targetTemp": 28.0}},
        {"roomId": "5", "type": "change_speed", "payload": {"speed": "HIGH"}},
    ],
    6: [
        {"roomId": "1", "type": "change_speed", "payload": {"speed": "HIGH"}},
    ],
    8: [
        {"roomId": "5", "type": "change_temp", "payload": {"targetTemp": 24.0}},
    ],
    10: [
        {"roomId": "1", "type": "change_temp", "payload": {"targetTemp": 22.0}},
        {"roomId": "4", "type": "change_temp", "payload": {"targetTemp": 21.0}},
        {"roomId": "4", "type": "change_speed", "payload": {"speed": "HIGH"}},
    ],
    12: [
        {"roomId": "5", "type": "change_speed", "payload": {"speed": "MID"}},
    ],
    13: [
        {"roomId": "2", "type": "change_speed", "payload": {"speed": "HIGH"}},
    ],
    15: [
        {"roomId": "1", "type": "power_off", "payload": {}},
        {"roomId": "3", "type": "change_speed", "payload": {"speed": "LOW"}},
    ],
    17: [
        {"roomId": "5", "type": "power_off", "payload": {}},
    ],
    18: [
         {"roomId": "3", "type": "change_speed", "payload": {"speed": "HIGH"}},
    ],
    19: [
        {"roomId": "1", "type": "power_on"},
        {"roomId": "4", "type": "change_temp", "payload": {"targetTemp": 25.0}},
        {"roomId": "4", "type": "change_speed", "payload": {"speed": "MID"}},
    ],
    21: [
        {"roomId": "2", "type": "change_temp", "payload": {"targetTemp": 26.0}},
        {"roomId": "2", "type": "change_speed", "payload": {"speed": "MID"}},
        {"roomId": "5", "type": "power_on"},
    ],
    25: [
        {"roomId": "1", "type": "power_off", "payload": {}},
        {"roomId": "3", "type": "power_off", "payload": {}},
        {"roomId": "5", "type": "power_off", "payload": {}},
    ],
    26: [
        {"roomId": "2", "type": "power_off", "payload": {}},
        {"roomId": "4", "type": "power_off", "payload": {}},
    ],
}

# ---------------------------------------------------------------------------
# 可选：在每分钟操作后抓一份监控视图，方便对照 Excel。
SNAPSHOT_ROOMS = {room["roomId"] for room in ROOM_PRESETS}


def load_config(path: Optional[str]) -> None:
    """Load external config to override baseUrl, hyperparams, rooms, and timeline.

    Supported formats: JSON (.json) and YAML (.yml/.yaml) if PyYAML is available.
    Structure:
    {
      "baseUrl": "http://localhost:8000",
      "hyperparams": { ... },
      "rooms": [ {"roomId": "room1", "initialTemp": 20.0, "ratePerNight": 100.0}, ... ],
      "timeline": { "1": [{"roomId": "room1", "type": "power_on", "payload": {"mode": "heat", "speed": "MID"}}], ... }
    }
    """
    global BASE_URL, HYPERPARAM_OVERRIDES, ROOM_PRESETS, TIMELINE
    if not path:
        return

    file = Path(path)
    if not file.exists():
        raise FileNotFoundError(f"Config file not found: {path}")

    def _coerce_timeline_keys(d: Dict[Any, Any]) -> Dict[int, List[Dict[str, Any]]]:
        result: Dict[int, List[Dict[str, Any]]] = {}
        for k, v in d.items():
            try:
                ik = int(k)
            except Exception:
                raise ValueError(f"Timeline minute keys must be integers: got {k}")
            if not isinstance(v, list):
                raise ValueError(f"Timeline minute {ik} must be a list of actions")
            result[ik] = v
        return result

    content: Dict[str, Any]
    if file.suffix.lower() in (".yml", ".yaml") and yaml is not None:
        content = yaml.safe_load(file.read_text(encoding="utf-8")) or {}
    else:
        content = json.loads(file.read_text(encoding="utf-8")) or {}

    if "baseUrl" in content and isinstance(content["baseUrl"], str):
        BASE_URL = content["baseUrl"].rstrip("/")
    if "hyperparams" in content and isinstance(content["hyperparams"], dict):
        HYPERPARAM_OVERRIDES = {**HYPERPARAM_OVERRIDES, **content["hyperparams"]}
    if "rooms" in content and isinstance(content["rooms"], list):
        ROOM_PRESETS = content["rooms"]
    if "timeline" in content and isinstance(content["timeline"], dict):
        TIMELINE = _coerce_timeline_keys(content["timeline"])
    # refresh snapshot set after overrides
    global SNAPSHOT_ROOMS
    SNAPSHOT_ROOMS = {room["roomId"] for room in ROOM_PRESETS}


def main() -> None:
    args = parse_args()
    load_config(args.config)
    global DRY_RUN
    DRY_RUN = bool(args.dry_run)
    if args.base_url:
        # Override via CLI if provided
        update_base_url(args.base_url)

    defaults = fetch_hyperparams()
    applied = configure_hyperparams(defaults)
    
    # 同步 tick_interval，确保后端时间推进与测试脚本一致
    clock_ratio = applied.get("clockRatio", 1.0)
    configure_tick_interval(clock_ratio)
    
    open_rooms(ROOM_PRESETS)
    check_in_rooms(ROOM_PRESETS)
    simulate_timeline(clock_ratio, max_minutes=args.max_minutes, step_by_step=args.step_by_step)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Replay AC control timeline per requirements")
    parser.add_argument("--config", type=str, default=None, help="Path to JSON/YAML config with requirements")
    parser.add_argument("--dry-run", action="store_true", help="Print requests without sending")
    parser.add_argument("--base-url", type=str, default=None, help="Override backend base URL (e.g. http://localhost:8000)")
    parser.add_argument("--max-minutes", type=int, default=None, help="Limit replay to N minutes")
    parser.add_argument("--step-by-step", action="store_true", help="Enable step-by-step debugging mode (pause system after each minute)")
    return parser.parse_args()


def update_base_url(url: str) -> None:
    global BASE_URL
    BASE_URL = url.rstrip("/")

# --- HTTP helpers ---------------------------------------------------------

def fetch_hyperparams() -> Dict[str, Any]:
    if DRY_RUN:
        print(f"[DRY] GET {BASE_URL}/monitor/hyperparams")
        # Return defaults so subsequent steps proceed
        return {
            "maxConcurrent": 3,
            "timeSliceSeconds": 60,
            "changeTempMs": 1000,
            "autoRestartThreshold": 0.5,
            "idleDriftPerMin": 0.3,
            "midDeltaPerMin": 0.5,
            "highMultiplier": 1.2,
            "lowMultiplier": 0.8,
            "defaultTarget": 23.0,
            "pricePerUnit": 1.0,
            "rateHighUnitPerMin": 1.0,
            "rateMidUnitPerMin": 0.5,
            "rateLowUnitPerMin": 1.0 / 3.0,
            "ratePerNight": 150.0,
            "clockRatio": 60.0,
        }
    resp = SESSION.get(f"{BASE_URL}/monitor/hyperparams", timeout=5)
    resp.raise_for_status()
    return resp.json()


def configure_hyperparams(defaults: Dict[str, Any]) -> Dict[str, Any]:
    payload = {
        key: (HYPERPARAM_OVERRIDES.get(key) if HYPERPARAM_OVERRIDES.get(key) is not None else defaults[key])
        for key in defaults
    }
    table = Table(title="Applying Hyperparameters", box=box.SIMPLE, show_header=False)
    for k in (
        "maxConcurrent","timeSliceSeconds","changeTempMs","autoRestartThreshold",
        "idleDriftPerMin","midDeltaPerMin","highMultiplier","lowMultiplier",
        "defaultTarget","pricePerUnit","rateHighUnitPerMin","rateMidUnitPerMin",
        "rateLowUnitPerMin","ratePerNight","clockRatio",
    ):
        table.add_row(f"[bold cyan]{k}[/]", str(payload.get(k)))
    CONSOLE.print(table)
    if DRY_RUN:
        CONSOLE.print(Panel.fit(f"[DRY] POST {BASE_URL}/monitor/hyperparams", title="Dry Run", border_style="magenta"))
        return payload
    resp = SESSION.post(f"{BASE_URL}/monitor/hyperparams", json=payload, timeout=5)
    resp.raise_for_status()
    applied = resp.json()
    CONSOLE.print("[green]✔ Hyperparameters applied[/]")
    return applied


def configure_tick_interval(clock_ratio: float) -> None:
    """根据 clockRatio 调整后端 tick 间隔，使时间加速与测试用例一致。"""
    if clock_ratio <= 0:
        return
    # 时钟倍率 = 相对正常速度的倍数；interval 越小越快
    interval = max(0.01, min(10.0, 1.0 / float(clock_ratio)))
    if DRY_RUN:
        CONSOLE.print(
            Panel.fit(
                f"[DRY] PUT {BASE_URL}/monitor/tick-interval\ninterval={interval}",
                title="Dry Run",
                border_style="magenta",
            )
        )
        return
    try:
        resp = SESSION.put(f"{BASE_URL}/monitor/tick-interval", json={"interval": interval}, timeout=5)
        resp.raise_for_status()
        body = resp.json()
        t = Table(title="Tick Interval", box=box.SIMPLE, show_header=False)
        t.add_row("interval", str(body.get("interval")))
        t.add_row("speedMultiplier", str(body.get("speedMultiplier")))
        CONSOLE.print(t)
    except requests.RequestException as exc:
        CONSOLE.print(f"[yellow]⚠ Failed to configure tick interval: {exc}[/]")


def open_rooms(presets: Iterable[Dict[str, Any]]) -> None:
    for room in presets:
        if DRY_RUN:
            CONSOLE.print(Panel.fit(f"[DRY] POST {BASE_URL}/monitor/rooms/open\nroom={room['roomId']}", title="Dry Run", border_style="magenta"))
            CONSOLE.print(f"[green]✔ Initialized room {room['roomId']} (idle)[/]")
            continue
        resp = SESSION.post(f"{BASE_URL}/monitor/rooms/open", json=room, timeout=5)
        resp.raise_for_status()
        CONSOLE.print(f"[green]✔ Initialized room {room['roomId']}[/]")


def check_in_rooms(presets: Iterable[Dict[str, Any]]) -> None:
    """Front-desk check-in for each opened room so it becomes OCCUPIED.

    Payload fields per API: custId, custName, guestCount, checkInDate, roomId, deposit
    """
    from datetime import datetime, timezone
    for idx, room in enumerate(presets, start=1):
        payload = {
            "custId": f"ID{room['roomId']}",
            "custName": f"Guest{idx}",
            "guestCount": 1,
            "checkInDate": datetime.now(timezone.utc).isoformat(),
            "roomId": room["roomId"],
            "deposit": 0.0,
        }
        if DRY_RUN:
            CONSOLE.print(Panel.fit(f"[DRY] POST {BASE_URL}/checkin\nroom={room['roomId']}", title="Dry Run", border_style="magenta"))
            CONSOLE.print(f"[green]✔ Checked-in (simulated) room {room['roomId']}[/]")
            continue
        resp = SESSION.post(f"{BASE_URL}/checkin", json=payload, timeout=5)
        resp.raise_for_status()
        body = resp.json()
        t = Table(title=f"Checked-in Room {room['roomId']}", box=box.SIMPLE, show_header=False)
        t.add_row("orderId", str(body.get("orderId")))
        t.add_row("status", str(body.get("status")))
        CONSOLE.print(t)

# --- Timeline execution ---------------------------------------------------

def pause_system() -> bool:
    """暂停系统（调试功能）"""
    if DRY_RUN:
        CONSOLE.print(Panel.fit(
            f"[DRY] POST {BASE_URL}/debug/system/pause",
            title="Dry Run",
            border_style="magenta"
        ))
        return True
    
    try:
        resp = SESSION.post(f"{BASE_URL}/debug/system/pause", timeout=5)
        resp.raise_for_status()
        result = resp.json()
        CONSOLE.print(Panel(
            f"[yellow]⏸️  系统已暂停[/]\n"
            f"Tick: {result.get('tick', 'N/A')}\n"
            f"{result.get('message', '')}",
            title="🛑 System Paused",
            border_style="yellow"
        ))
        return True
    except requests.RequestException as exc:
        CONSOLE.print(Panel(
            f"[red]⚠ 暂停系统失败:[/]\n{exc}",
            title="Error",
            border_style="red"
        ))
        return False


def resume_system() -> bool:
    """恢复系统（调试功能）"""
    if DRY_RUN:
        CONSOLE.print(Panel.fit(
            f"[DRY] POST {BASE_URL}/debug/system/resume",
            title="Dry Run",
            border_style="magenta"
        ))
        return True
    
    try:
        resp = SESSION.post(f"{BASE_URL}/debug/system/resume", timeout=5)
        resp.raise_for_status()
        result = resp.json()
        CONSOLE.print(Panel(
            f"[green]▶️  系统已恢复[/]\n"
            f"Tick: {result.get('tick', 'N/A')}\n"
            f"{result.get('message', '')}",
            title="✅ System Resumed",
            border_style="green"
        ))
        return True
    except requests.RequestException as exc:
        CONSOLE.print(Panel(
            f"[red]⚠ 恢复系统失败:[/]\n{exc}",
            title="Error",
            border_style="red"
        ))
        return False


def simulate_timeline(clock_ratio: float, max_minutes: Optional[int] = None, step_by_step: bool = False) -> None:
    minute_step = 60.0 / max(clock_ratio, 0.01)
    max_minute = max(TIMELINE.keys(), default=0)
    if max_minutes is not None:
        max_minute = min(max_minute, max_minutes)
    
    # 当前时钟倍率（单步调试模式下可动态调整）
    current_clock_ratio = clock_ratio
    
    CONSOLE.print(Panel.fit(
        f"minutes={max_minute}\nclockRatio={clock_ratio}\nminute_step={minute_step:.2f}s\nDRY_RUN={DRY_RUN}\nSTEP_BY_STEP={step_by_step}", 
        title="Starting Timeline", 
        border_style="cyan"
    ))

    # 等待后端完全处理 tick_interval 配置后再开始
    if not DRY_RUN:
        CONSOLE.print("[yellow]等待 2 秒让后端同步 tick_interval...[/]")
        time.sleep(2)
    else:
        CONSOLE.print("[yellow]DRY_RUN 模式，跳过等待[/]")

    for minute in range(0, max_minute + 1):
        actions = TIMELINE.get(minute, [])
        
        # 执行该分钟的操作
        if actions:
            CONSOLE.print(Panel.fit(f"Minute {minute}", border_style="blue"))
            for action in actions:
                send_action(action)
        else:
            CONSOLE.print(f"[dim]Minute {minute}: No actions[/]")
        
        # 使用时钟同步+快照接口，每分钟都等待 60 个 tick 完成（1 分钟业务时间）并在 tick 线程中立即采集快照
        if not DRY_RUN:
            # 计算基于当前时钟倍率的 tick 间隔
            tick_interval = 60.0 / max(current_clock_ratio, 0.01) / 60  # 计算每个 tick 的时间
            expected_time = 60 * tick_interval
            # 超时时间设置为预期时间的 20 倍，确保即使 CPU 负载很高也不会超时
            timeout = max(30.0, expected_time * 20)
            
            info_panel = Panel(
                f"[cyan]分钟 {minute}: 等待 60 个 tick 完成并在 tick 线程中采集快照[/]\n"
                f"预计耗时: [yellow]{expected_time:.2f}[/] 秒\n"
                f"超时设置: [yellow]{timeout:.1f}[/] 秒\n"
                f"机制: [green]Snapshot in tick thread (blocks tick)[/]\n"
                f"DRY_RUN: [red]{DRY_RUN}[/]",
                title="⏱️ Time Sync + Snapshot (Blocking)",
                border_style="cyan"
            )
            CONSOLE.print(info_panel)
            
            # 只有在有操作或第 0 分钟时才采集快照
            if actions or minute == 0:
                if not wait_for_tick_and_snapshot(minute=minute, count=60, timeout=timeout):
                    CONSOLE.print(Panel(
                        "[red]⚠ 时钟同步超时，使用 sleep 备用方案[/]",
                        border_style="red"
                    ))
                    time.sleep(minute_step)
                    # 使用旧的快照接口作为备用
                    snapshot_rooms(minute)
            else:
                # 没有操作时只等待，不采集快照
                if not wait_for_tick_and_snapshot(minute=minute, count=60, timeout=timeout):
                    CONSOLE.print(Panel(
                        "[red]⚠ 时钟同步超时，使用 sleep 备用方案[/]",
                        border_style="red"
                    ))
                    time.sleep(minute_step)
            
            # 单步调试模式：每分钟后暂停系统，等待用户确认
            if step_by_step:
                pause_system()
                
                # 显示当前状态和可用命令
                CONSOLE.print(Panel(
                    f"[cyan]📍 已完成分钟 {minute}[/]\n"
                    f"[yellow]系统已暂停，可以查看调试管理员界面检查状态[/]\n\n"
                    f"[bold]当前时钟倍率:[/] [green]{current_clock_ratio}x[/] (1分钟 ≈ {60.0/max(current_clock_ratio, 0.01):.2f}秒)\n\n"
                    f"[bold]可用命令:[/]\n"
                    f"  [green]Enter[/]          - 继续下一分钟\n"
                    f"  [cyan]speed <ratio>[/]  - 调整时钟倍率 (例如: speed 120)\n"
                    f"  [magenta]info[/]           - 显示当前配置\n"
                    f"  [red]q[/]              - 退出测试",
                    title="⏸️  Step-by-Step Debug Mode",
                    border_style="cyan"
                ))
                
                while True:
                    user_input = input("> ").strip()
                    
                    if user_input.lower() == 'q':
                        CONSOLE.print("[yellow]用户中止测试[/]")
                        return
                    elif user_input.lower() == 'info':
                        # 显示当前配置信息
                        info_table = Table(title="当前配置", box=box.SIMPLE, show_header=False)
                        info_table.add_row("当前分钟", str(minute))
                        info_table.add_row("总分钟数", str(max_minute))
                        info_table.add_row("时钟倍率", f"{current_clock_ratio}x")
                        info_table.add_row("1分钟耗时", f"{60.0/max(current_clock_ratio, 0.01):.2f}秒")
                        info_table.add_row("Tick间隔", f"{1.0/current_clock_ratio:.4f}秒")
                        CONSOLE.print(info_table)
                    elif user_input.lower().startswith('speed '):
                        # 调整时钟倍率
                        try:
                            parts = user_input.split()
                            new_ratio = float(parts[1])
                            if new_ratio <= 0:
                                CONSOLE.print("[red]❌ 时钟倍率必须大于 0[/]")
                                continue
                            if new_ratio > 1000:
                                CONSOLE.print("[yellow]⚠ 时钟倍率过高可能导致系统不稳定，建议使用 <= 1000[/]")
                            
                            # 更新时钟倍率
                            current_clock_ratio = new_ratio
                            configure_tick_interval(current_clock_ratio)
                            
                            CONSOLE.print(Panel(
                                f"[green]✅ 时钟倍率已调整为 {current_clock_ratio}x[/]\n"
                                f"1分钟业务时间 ≈ {60.0/max(current_clock_ratio, 0.01):.2f}秒真实时间\n"
                                f"Tick间隔: {1.0/current_clock_ratio:.4f}秒",
                                title="⚡ Speed Updated",
                                border_style="green"
                            ))
                        except (ValueError, IndexError):
                            CONSOLE.print("[red]❌ 无效的命令格式。使用: speed <数字>[/]")
                    elif user_input == '':
                        # 按 Enter 继续
                        break
                    else:
                        CONSOLE.print("[yellow]⚠ 未知命令。可用命令: Enter, speed <ratio>, info, q[/]")
                
                # 恢复系统继续
                resume_system()
        else:
            CONSOLE.print(Panel(
                f"[yellow]DRY_RUN 模式: 跳过 wait_for_tick_and_snapshot (minute={minute})[/]",
                border_style="yellow"
            ))

    CONSOLE.print("[green]✔ Timeline replay finished[/]")
    export_excel_snapshots(SNAPSHOT_ROWS)


def send_action(action: Dict[str, Any]) -> None:
    room_id = action["roomId"]
    action_type = action["type"]
    payload = action.get("payload") or {}

    if action_type == "power_on":
        path = f"/rooms/{room_id}/ac/power-on"
    elif action_type == "power_off":
        path = f"/rooms/{room_id}/ac/power-off"
    elif action_type == "change_temp":
        path = f"/rooms/{room_id}/ac/change-temp"
    elif action_type == "change_speed":
        path = f"/rooms/{room_id}/ac/change-speed"
    else:
        raise ValueError(f"Unknown action type: {action_type}")

    if DRY_RUN:
        CONSOLE.print(Panel.fit(f"[DRY] POST {BASE_URL}{path}", title="Dry Run", border_style="magenta"))
        return
    
    try:
        resp = SESSION.post(f"{BASE_URL}{path}", json=payload if payload else None, timeout=5)
        resp.raise_for_status()
        body = resp.json()
        t = Table(title=f"{action_type.upper()} → Room {room_id}", box=box.SIMPLE, show_header=False)
        t.add_row("status", str(body.get("status")))
        t.add_row("isServing", str(body.get("isServing")))
        t.add_row("isWaiting", str(body.get("isWaiting")))
        CONSOLE.print(t)
    except requests.HTTPError as e:
        # 捕获 HTTP 错误（如 400 Bad Request），显示错误信息但不中断测试
        error_detail = "Unknown error"
        if e.response is not None:
            try:
                error_body = e.response.json()
                error_detail = error_body.get("detail", str(e))
            except:
                error_detail = e.response.text or str(e)
        
        error_panel = Panel(
            f"[red]❌ {action_type.upper()} → Room {room_id} FAILED[/]\n"
            f"[yellow]{error_detail}[/]",
            title="⚠️ Request Rejected",
            border_style="red"
        )
        CONSOLE.print(error_panel)
    except requests.RequestException as e:
        # 捕获其他网络错误
        error_panel = Panel(
            f"[red]❌ {action_type.upper()} → Room {room_id} FAILED[/]\n"
            f"[yellow]{str(e)}[/]",
            title="⚠️ Network Error",
            border_style="red"
        )
        CONSOLE.print(error_panel)


def wait_for_tick_and_snapshot(minute: int, count: int = 1, timeout: float = 5.0) -> bool:
    """
    等待指定数量的 tick 完成并在 tick 线程中立即采集快照(阻塞 tick)
    
    通过在 tick 线程中同步执行快照采集,确保快照时间戳与 tick 推进完全一致,
    完全消除了异步等待和快照采集之间可能产生的额外 tick 导致的时间偏移。
    
    参数:
    - minute: 当前分钟数(用于显示)
    - count: 要等待的 tick 数量
    - timeout: 总超时时间(秒)

    返回 True 表示成功，False 表示超时。
    """
    if DRY_RUN:
        CONSOLE.print(Panel.fit(
            f"[DRY] POST {BASE_URL}/monitor/wait-tick-and-snapshot\ncount={count}, timeout={timeout:.1f}s",
            title="Dry Run",
            border_style="magenta"
        ))
        return True

    try:
        url = f"{BASE_URL}/monitor/wait-tick-and-snapshot"
        params = {"count": count, "timeout": timeout}

        # 使用 Rich Table 显示调用信息
        t = Table(title="🕑 Waiting for Tick + Snapshot (Blocking)", box=box.SIMPLE, show_header=False)
        t.add_row("URL", f"{url}")
        t.add_row("count", str(count))
        t.add_row("timeout", f"{timeout:.1f}s")
        t.add_row("mechanism", "[cyan]Snapshot in tick thread (blocks next tick)[/]")
        CONSOLE.print(t)

        resp = SESSION.post(url, params=params, timeout=timeout + 1)
        resp.raise_for_status()
        result = resp.json()

        success = result.get("success", False)
        tick_counter = result.get("tickCounter", 0)
        message = result.get("message", "")
        snapshot = result.get("snapshot")

        # 显示结果
        result_table = Table(
            title="✅ Tick Sync + Snapshot Result" if success else "⚠️ Tick Sync Failed",
            box=box.SIMPLE,
            show_header=False
        )
        result_table.add_row("success", str(success))
        result_table.add_row("tickCounter", str(tick_counter))
        result_table.add_row("message", message)
        if success:
            result_table.add_row("mechanism", "[green]✓ Snapshot captured in tick thread[/]")
        CONSOLE.print(result_table)

        # 处理快照数据
        if success and snapshot:
            rooms = snapshot.get("rooms", [])
            for room in rooms:
                if room["roomId"] in SNAPSHOT_ROOMS:
                        SNAPSHOT_ROWS.append({
                        "minute": minute,
                        "roomId": room["roomId"],
                        "status": room["status"],
                        "currentTemp": float(room["currentTemp"]),
                        "targetTemp": float(room["targetTemp"]),
                        "speed": room["speed"] or "",
                        "currentFee": float(room.get("currentFee", 0.0)),
                        "totalFee": float(room.get("totalFee", 0.0)),
                        "servedSeconds": int(room.get("servedSeconds", 0)),
                        "waitedSeconds": int(room.get("waitedSeconds", 0)),
                        "isServing": bool(room.get("isServing")),
                        "isWaiting": bool(room.get("isWaiting")),
                    })

            CONSOLE.print(f"[green]✔ Snapshot captured for minute {minute}[/]")

        return success
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            CONSOLE.print(Panel(
                f"[yellow]⚠ API not found: {url}[/]\nFalling back to sleep-based time progression.",
                title="API Not Found",
                border_style="yellow"
            ))
        else:
            CONSOLE.print(Panel(
                f"[red]⚠ Wait for tick and snapshot 调用失败:[/]\n{exc}",
                title="Error",
                border_style="red"
            ))
        return False
    except requests.RequestException as exc:
        CONSOLE.print(Panel(
            f"[red]⚠ Wait for tick and snapshot 调用失败:[/]\n{exc}",
            title="Error",
            border_style="red"
        ))
        return False


def snapshot_rooms(minute: int) -> None:
    try:
        if DRY_RUN:
            CONSOLE.print(Panel.fit(f"[DRY] GET {BASE_URL}/monitor/rooms", title="Dry Run", border_style="magenta"))
            return
        resp = SESSION.get(f"{BASE_URL}/monitor/rooms", timeout=5)
        resp.raise_for_status()
    except requests.RequestException as exc:
        CONSOLE.print(f"[yellow]⚠ Snapshot failed at minute {minute}: {exc}[/]")
        return

    rooms = resp.json().get("rooms", [])
    # 将监控接口返回的队列信息一并采集（servedSeconds / waitedSeconds / isServing / isWaiting）
    summary = [
        {
            "roomId": room["roomId"],
            "status": room["status"],
            "currentTemp": room["currentTemp"],
            "targetTemp": room["targetTemp"],
            "speed": room["speed"],
            "currentFee": round(room.get("currentFee", 0.0), 2),
            "totalFee": round(room.get("totalFee", 0.0), 2),
            "servedSeconds": int(room.get("servedSeconds", 0)),
            "waitedSeconds": int(room.get("waitedSeconds", 0)),
            "isServing": bool(room.get("isServing")),
            "isWaiting": bool(room.get("isWaiting")),
        }
        for room in rooms
        if room["roomId"] in SNAPSHOT_ROOMS
    ]
    if summary:
        table = Table(title=f"Snapshot @ minute {minute}", box=box.SIMPLE)
        table.add_column("Room")
        table.add_column("Status")
        table.add_column("Temp")
        table.add_column("Speed")
        table.add_column("Session Fee")
        table.add_column("Total Fee")
        for r in summary:
            table.add_row(
                str(r["roomId"]),
                str(r["status"]),
                f"{r['currentTemp']:.1f}℃ → {r['targetTemp']:.1f}℃",
                str(r["speed"]),
                f"¥{r['currentFee']:.2f}",
                f"¥{r['totalFee']:.2f}",
            )
            # accumulate raw rows for Excel export（连同队列统计信息）
            SNAPSHOT_ROWS.append({
                "minute": minute,
                "roomId": r["roomId"],
                "status": r["status"],
                "currentTemp": float(r["currentTemp"]),
                "targetTemp": float(r["targetTemp"]),
                "speed": r["speed"] or "",
                "currentFee": float(r["currentFee"]),
                "totalFee": float(r["totalFee"]),
                "servedSeconds": int(r.get("servedSeconds", 0)),
                "waitedSeconds": int(r.get("waitedSeconds", 0)),
                "isServing": bool(r.get("isServing")),
                "isWaiting": bool(r.get("isWaiting")),
            })
        CONSOLE.print(table)


def export_excel_snapshots(rows: List[Dict[str, Any]], filename: str = "snapshot_report.xlsx") -> None:
    if not rows:
        CONSOLE.print("[yellow]⚠ No snapshots to export[/]")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "制热测试用例"

    # styles（简化配色，只保留头部底色，数据区不再按状态上色）
    header_fill = PatternFill("solid", fgColor="FFF2CC")
    subheader_fill = PatternFill("solid", fgColor="FFF2CC")

    # Build unique minutes and rooms
    minutes = sorted({r["minute"] for r in rows})
    rooms = [str(r) for r in sorted({int(str(x["roomId"])) for x in rows})]

    # Header rows: Row1 has room group titles, Row2 has subheaders per room
    ws.append(["时间(min)"] + [None] * (len(rooms) * 4) + ["服务队列", "等待队列"])  # row 1 placeholder
    ws.append(["时间(min)"] + sum([["房间" + room, "当前", "目标", "风速", "费用"] for room in rooms], [])[:len(rooms)*4+len(rooms)] + ["服务队列", "等待队列"])  # row 2 human labels

    # Actually set merged headers with colors
    ws.cell(row=1, column=1, value="时间(min)")
    ws.cell(row=2, column=1, value="时间(min)")
    ws.cell(row=1, column=len(rooms)*4 + 2, value="服务队列")
    ws.cell(row=1, column=len(rooms)*4 + 3, value="等待队列")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=len(rooms)*4 + 2, end_row=2, end_column=len(rooms)*4 + 2)
    ws.merge_cells(start_row=1, start_column=len(rooms)*4 + 3, end_row=2, end_column=len(rooms)*4 + 3)

    # Create room group headers and subheaders
    col = 2
    for room in rooms:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        ws.cell(row=1, column=col, value=f"房间{room}")
        ws.cell(row=2, column=col, value="当前")
        ws.cell(row=2, column=col + 1, value="目标")
        ws.cell(row=2, column=col + 2, value="风速")
        ws.cell(row=2, column=col + 3, value="费用")
        for c in range(col, col + 4):
            ws.cell(row=1, column=c).fill = header_fill
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=c).fill = subheader_fill
            ws.cell(row=2, column=c).font = Font(bold=True)
            ws.cell(row=2, column=c).alignment = Alignment(horizontal="center")
        col += 4

    # Style time and queue headers
    for r in (1, 2):
        ws.cell(row=r, column=1).fill = header_fill
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=len(rooms)*4 + 2).fill = header_fill
        ws.cell(row=r, column=len(rooms)*4 + 2).font = Font(bold=True)
        ws.cell(row=r, column=len(rooms)*4 + 2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=len(rooms)*4 + 3).fill = header_fill
        ws.cell(row=r, column=len(rooms)*4 + 3).font = Font(bold=True)
        ws.cell(row=r, column=len(rooms)*4 + 3).alignment = Alignment(horizontal="center")

    # Index rows by minute+room and prepare forward-fill per room
    data = {}
    for r in rows:
        key = (r["minute"], str(r["roomId"]))
        data[key] = r
    min_minute = min(minutes)
    max_minute = max(minutes)
    minutes_full = list(range(min_minute, max_minute + 1))
    last_by_room = {room: None for room in rooms}

    # Fill rows per minute
    current_row = 3
    for m in minutes_full:
        ws.cell(row=current_row, column=1, value=m)
        col = 2
        # 构建本分钟的队列字符串：房间ID/时间（秒），仅展示在一行的“服务队列/等待队列”列
        minute_rows = [r for r in rows if r["minute"] == m]
        serving_pairs = []
        waiting_pairs = []
        for r in minute_rows:
            if r.get("isServing"):
                serving_pairs.append(f"R{r['roomId']}/{int(r.get('servedSeconds', 0))}")
            if r.get("isWaiting"):
                waiting_pairs.append(f"R{r['roomId']}/{int(r.get('waitedSeconds', 0))}")
        service_str = " ".join(sorted(serving_pairs)) if serving_pairs else ""
        wait_str = " ".join(sorted(waiting_pairs)) if waiting_pairs else ""

        for room in rooms:
            r = data.get((m, room))
            if r:
                ws.cell(row=current_row, column=col, value=round(r["currentTemp"], 1))
                ws.cell(row=current_row, column=col + 1, value=round(r["targetTemp"], 1))
                ws.cell(row=current_row, column=col + 2, value=r["speed"])
                # 费用列使用累计费用（totalFee）
                ws.cell(row=current_row, column=col + 3, value=round(r["totalFee"], 2))
                for c in range(col, col + 4):
                    ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")
                last_by_room[room] = r
            else:
                # forward-fill with last known snapshot for this room if available
                lr = last_by_room.get(room)
                if lr:
                    ws.cell(row=current_row, column=col, value=round(lr["currentTemp"], 1))
                    ws.cell(row=current_row, column=col + 1, value=round(lr["targetTemp"], 1))
                    ws.cell(row=current_row, column=col + 2, value=lr["speed"])
                    ws.cell(row=current_row, column=col + 3, value=round(lr["totalFee"], 2))
                    for c in range(col, col + 4):
                        ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")
                else:
                    for c in range(col, col + 4):
                        ws.cell(row=current_row, column=c, value=None)
                        ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")
            col += 4
        # queue columns：使用当前分钟的服务/等待队列摘要
        ws.cell(row=current_row, column=len(rooms)*4 + 2, value=service_str)
        ws.cell(row=current_row, column=len(rooms)*4 + 3, value=wait_str)
        current_row += 1

    # auto-width (avoid MergedCell by using column index)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(9, min(18, max_len + 2))

    try:
        wb.save(filename)
        CONSOLE.print(f"[green]✔ Excel exported: {filename}[/]")
    except Exception as exc:
        CONSOLE.print(f"[red]Failed to write Excel: {exc}[/]")


if __name__ == "__main__":
    main()
